from openpyxl import Workbook,load_workbook
import sys
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QDialog
import numpy as np
import ast
class PlotWindow(QDialog):
    def __init__(self,variables):
        super().__init__()
        self.setWindowTitle("Scatter PLot")
        self.setGeometry(500,500,500,500)
        self.canvas=FigureCanvas(plt.figure())
        layout=QVBoxLayout()
        layout.addWidget(self.canvas)
        self.plot_graph(variables)
        self.setLayout(layout)
    def plot_graph(self,variables):
        try:
            import ast
            import matplotlib.pyplot as plt
            import numpy as np
            x_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[0]))
            y_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[1]))
            x_axis=np.array(x_axis)
            y_axis=np.array(y_axis)
            x_label=variables[2]
            y_label=variables[3]
            fig,ax=plt.subplots(figsize=(6,4))
            ax.scatter(x_axis,y_axis,color="blue",alpha=0.7)
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.set_title("Scatter Plot")
            ax.grid()
            self.canvas.figure=fig
            self.canvas.draw
        except Exception as e:
            print(str(e))

def retrieve_from_excel(path):
    alphabet_dict = {chr(letter): index for index, letter in enumerate(range(ord('A'), ord('Z') + 1), start=1)}
    alphabets = [chr(i) for i in range(65, 91)]
    numbers = [i for i in range(10)]
    wb=load_workbook(path,data_only=True)
    ws=wb.active
    i=1
    q=1
    if ws.cell(row=1,column=1).value is None:
        html="<H3 style='color:grey;background-color:#2e2b2b;'>No values to be shown</H3>"
        return html
    while  (ws.cell(row=i,column=q).value):
        i=i+1
    row=i-1
    i=1
    while  (ws.cell(row=i,column=q).value):
        q=q+1
    column=q-1
    q=1
    html=""""""
    html=html+"<table style='background-color:white;color:black;border-collapse:collapse;'>"
    for i in range(1,row+1):
        html=html+"<tr>"
        for q in range(1,column+1):
            html=html+f"<td style='border: 5px solid black; padding: 5px;'>{ws.cell(row=i,column=q).value}</td>"
        html=html+"</tr>"
    return(html)
def retrieve_from_excel_console(path):
    alphabet_dict = {chr(letter): index for index, letter in enumerate(range(ord('A'), ord('Z') + 1), start=1)}
    alphabets = [chr(i) for i in range(65, 91)]
    numbers = [i for i in range(10)]
    wb=load_workbook(path,data_only=True)
    ws=wb.active
    i=1
    q=1
    if ws.cell(row=1,column=1).value is None:
        html="<H3 style='color:white;background-color:#3b3847;'>No values to be shown</H3>"
        return html
    while  (ws.cell(row=i,column=q).value):
        i=i+1
    row=i-1
    i=1
    while  (ws.cell(row=i,column=q).value):
        q=q+1
    column=q-1
    q=1
    html=""""""
    html=html+"<table style='margin-left:10px;margin-top:20px;margin-bottom:30px;background-color:#3b3847;color:white;' border='0'>"
    for i in range(1,row+1):
        html=html+"<tr>"
        for q in range(1,column+1):
            html=html+f"<td style='background-color:#3b3847;color:white;padding: 5px;'>{ws.cell(row=i,column=q).value}</td>"
        html=html+"</tr>"
    return(html)
def retrieve_from_excel_variable_values(path):
    alphabet_dict = {chr(letter): index for index, letter in enumerate(range(ord('A'), ord('Z') + 1), start=1)}
    alphabets = [chr(i) for i in range(65, 91)]
    numbers = [i for i in range(10)]
    wb=load_workbook(path,data_only=True)
    ws=wb.active
    i=1
    q=1
    if ws.cell(row=1,column=1).value is None:
        html="<H3 style='color:grey;background-color:#2e2b2b;'>No values to be shown</H3>"
        return html
    while  (ws.cell(row=i,column=q).value):
        i=i+1
    row=i-1
    i=1
    while  (ws.cell(row=i,column=q).value):
        q=q+1
    column=q-1
    q=1
    html=""""""
    html=html+"<!DOCTYPE html><html><body style='width:100%;background-color:#2e2b2b;'><table width='100%' style='table-layout: fixed;background-color:#2e2b2b;color:white;border:2px solid grey;border-collapse:collapse;'>"
    for i in range(1,row+1):
        html=html+"<tr>"
        for q in range(1,column+1):
            html=html+f"<td style='border: 2px solid grey; padding: 5px;'>{ws.cell(row=i,column=q).value}</td>"
        html=html+"</tr>"
    html=html+"</table></body></html>"
    return(html)
def replace_excel(name,value_1):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    wb=load_workbook(file_path);
    ws=wb.active;i=1;q=0;
    print("Happening")
    while ws['A'+str(i)].value!=None:
        if ws['A'+str(i)].value==name:
            ws['B'+str(i)].value=value_1
            wb.save(file_path)
            break;
        i=i+1
import openpyxl
def column_retrieve(path,column):
    wb=openpyxl.load_workbook(path)
    ws=wb.active
    i=1
    while ws.cell(row=1,column=i) is not None:
        if ws.cell(row=1,column=i).value==column:
            break
        i=i+1
    result=[]
    j=2
    while ws.cell(row=j,column=i).value is not None:
        result.append(ws.cell(row=j,column=i).value)
        j=j+1
    return result
    
def check_excel_variable(name):
    print("check_excel_variable")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    wb=load_workbook(file_path);
    ws=wb.active;i=1;
    if ws['A'+str(i)]!=None:
        print("")
    while ws['A'+str(i)].value!=None:
        if ws['A'+str(i)].value==name:
            return True
        i=i+1
    return False
def tokenize_expression(expr):
    operators = {'+', '-', '*', '/', '?>', '<?', '??', '!?', '>', '<'}
    tokens = []
    current_token = ""
    i = 0

    while i < len(expr):
        # Check for two-character operators first
        if i < len(expr) - 1 and expr[i:i+2] in operators:
            if current_token:
                tokens.append(current_token)
                current_token = ""
            tokens.append(expr[i:i+2])
            i += 2  # Move two steps ahead
        elif expr[i] in operators:
            if current_token:
                tokens.append(current_token)
                current_token = ""
            tokens.append(expr[i])
            i += 1  # Move one step ahead
        else:
            current_token += expr[i]
            i += 1

    if current_token:
        tokens.append(current_token)

    return tokens

def cmd_retrieve_excel(name):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "Variable_Storage.xlsx")
    wb = load_workbook(file_path)
    ws = wb.active
    i = 1

    print(f"Retrieving {name} from Excel...")

    while ws['A' + str(i)].value is not None:
        if ws['A' + str(i)].value.strip() == name.strip():
            retrieved_value = ws['B' + str(i)].value
            wb.close()
            print(f"Found {name}: {retrieved_value}")
            return retrieved_value  # Return the found value
        i += 1

    wb.close()
    print(f"Variable '{name}' not found.")
    return "#ERROR"  # If not found, return error

from PyQt5.QtWidgets import QApplication,QPlainTextEdit,QTextEdit
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QTextCharFormat,QColor
import os
class CustomPlainTextEdit(QTextEdit):
    def __init__(self,addition):
        super().__init__()
        script_dir = os.path.dirname(os.path.abspath(__file__))
        credits_file_path = os.path.join(script_dir, 'Credits_LaplaceLab.txt')
        with open(credits_file_path,'r') as file:
            content=file.read()
            self.protected_text=content
            self.setText(content)
        self.setProtectedTextStyle()
        self.setReadOnly(True)                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                  

    def setProtectedTextStyle(self):
        cursor=self.textCursor()
        cursor.movePosition(cursor.Right,cursor.KeepAnchor,len(self.protected_text))
        format=QTextCharFormat()
        format.setForeground(QColor("grey"))
        format.setFontFamily("Arial")
        format.setFontPointSize(12)
        cursor.setCharFormat(format)
        

from tkinter import *
import math

def draw_points_on_canvas(canvas):
    canvas.delete("all")
    real = float(d.get())
    i = 0
    while i < 501:
        if i == 0:
            canvas.create_line(0, 500 - i, 1920, 500 - i, fill="red", width=4)
        else:
            canvas.create_line(0, 500 - i, 1920, 500 - i, fill="yellow", width=1)
        i += 80

    i = 0
    while i < 501:
        if i == 0:
            canvas.create_line(0, 500 + i, 1920, 500 + i, fill="red", width=4)
        else:
            canvas.create_line(0, 500 + i, 1920, 500 + i, fill="yellow", width=1)
        i += 80

    i = 0
    while i < 951:
        if i == 0:
            canvas.create_line(950 - i, 0, 950 - i, 1080, fill="red", width=4)
        else:
            canvas.create_line(950 - i, 0, 950 - i, 1080, fill="yellow", width=1)
        i += 80

    i = 0
    while i < 951:
        if i == 0:
            canvas.create_line(950 + i, 0, 950 + i, 1080, fill="red", width=4)
        else:
            canvas.create_line(950 + i, 0, 950 + i, 1080, fill="yellow", width=1)
        i += 80

    canvas.create_text(950, 500, text="(0,0)", font=("helvetica", 12, "bold"), fill="white")

    # Plot points on the x-axis
    j = real
    a = 950 - j
    i = -1
    while a > 0:
        canvas.create_oval(a - 5, 500 - 5, a + 5, 500 + 5, fill="white")
        canvas.create_text(a, 500 - 25, text=f"({i},0)", font=("helvetica", 12, "bold"), fill="white")
        i -= 1
        a -= j

    a = 950 + j
    i = 1
    while a < 1920:
        canvas.create_oval(a - 5, 500 - 5, a + 5, 500 + 5, fill="white")
        canvas.create_text(a, 500 - 25, text=f"({i},0)", font=("helvetica", 12, "bold"), fill="white")
        i += 1
        a += j

    # Plot points on the y-axis
    a = 500 + j
    i = 1
    while a < 1080:
        canvas.create_oval(950 - 5, a - 5, 950 + 5, a + 5, fill="white")
        canvas.create_text(990, a + 5, text=f"(-{i},0)", font=("helvetica", 12, "bold"), fill="white")
        i += 1
        a += j

    a = 500 - j
    i = 1
    while a > 0:
        canvas.create_oval(950 - 5, a - 5, 950 + 5, a + 5, fill="white")
        canvas.create_text(990, a + 5, text=f"({i},0)", font=("helvetica", 12, "bold"), fill="white")
        i += 1
        a -= j

    coefficient_x = float(x_coef.get())
    coefficient_y = float(y_coef.get())
    powerx = float(power_equation_x.get())
    powery = float(power_equation_y.get())
    constan = float(constant.get())
    
    x = 0
    css = 0
    while x < 1920:
        try:
            y = ((coefficient_x * x ** powerx) ** (1 / powery)) + constan
            canvas.create_oval(950 + (x * real) - 5, 500 - y * real - 5, 
                               950 + (x * real) + 5, 500 - y * real + 5, fill="red")

            if css % 15 == 0:
                m = "{:.2f}".format(x)
                n = "{:.2f}".format(y)
                canvas.create_text(950 + (x * real), 500 - y * real + 40, 
                                  text=f"({m},{n})", font=("helvetica", 12, "bold"), fill="#00FFFF")

            if x > 0:
                canvas.create_line(950 + (x1 * real), 500 - y1 * real, 
                                  950 + (x * real), 500 - y * real, fill="#FF00FF", width=3)

            x1, y1 = x, y
            css += 1
            x += 0.1
        except ValueError:
            break

    x = 0
    css = 0
    while x > -960:
        try:
            y = ((coefficient_x * x ** powerx) ** (1 / powery)) + constan
            canvas.create_oval(950 + (x * real) - 5, 500 - y * real - 5, 
                               950 + (x * real) + 5, 500 - y * real + 5, fill="red")

            if css % 15 == 0 and x != 0:
                m = "{:.2f}".format(x)
                n = "{:.2f}".format(y)
                canvas.create_text(950 + (x * real), 500 - y * real + 60, 
                                  text=f"({m},{n})", font=("helvetica", 12, "bold"), fill="#00FFFF")

            if x < 0:
                canvas.create_line(950 + (x1 * real), 500 - y1 * real, 
                                  950 + (x * real), 500 - y * real, fill="#FF00FF", width=3)

            x1, y1 = x, y
            css += 1
            x -= 0.1
        except ValueError:
            break

def plotter():
    global x_coef, y_coef, power_equation_x, power_equation_y, d, constant

    c = Tk()
    screen_width = c.winfo_screenwidth()
    screen_height = c.winfo_screenheight()
    c.geometry(f"{screen_width}x{screen_height}+0+0")

    canvas = Canvas(c, width=screen_width, height=screen_height, bg="black")
    canvas.pack()

    input1 = Toplevel(c)
    input1.geometry("600x200+0+0")

    Label(input1, text="power_x", fg="black").grid(row=0, column=0)
    power_equation_x = Entry(input1, width=10)
    power_equation_x.grid(row=0, column=1)

    Label(input1, text="power_y", fg="black").grid(row=1, column=0)
    power_equation_y = Entry(input1, width=10)
    power_equation_y.grid(row=1, column=1)

    Label(input1, text="Coeff.(x)", fg="black").grid(row=2, column=0)
    x_coef = Entry(input1, width=10)
    x_coef.grid(row=2, column=1)

    Label(input1, text="Coeff.(y)", fg="black").grid(row=3, column=0)
    y_coef = Entry(input1, width=10)
    y_coef.grid(row=3, column=1)

    Label(input1, text="dist", fg="black").grid(row=4, column=0)
    d = Entry(input1, width=10)
    d.grid(row=4, column=1)

    Label(input1, text="constant", fg="black").grid(row=5, column=0)
    constant = Entry(input1, width=10)
    constant.grid(row=5, column=1)

    Button(input1, text="Enter", command=lambda: draw_points_on_canvas(canvas)).grid(row=6, column=0)

    c.mainloop()


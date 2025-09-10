from functools import partial
from PyQt5.QtWidgets import (
    QSizePolicy, QTextEdit, QFrame, QTabWidget, QFileDialog, QWidget,
    QPlainTextEdit, QDesktopWidget, QAction, QApplication, QMainWindow,
    QPushButton, QVBoxLayout, QHBoxLayout
)
from PyQt5.QtCore import Qt, QRegularExpression
from PyQt5.QtGui import QColor, QTextCharFormat, QSyntaxHighlighter

import Laplacelab_lastcode as L
import threading
import os
import sys
import re

operators=['+', '-', '*', '/']    
console_log=''
s='''
QScrollBar:vertical{
boder-radius:6px;
}
QTabBar::scroller::add-page, QTabBar::scroller::sub-page {
    background: transparent;
    width: 0px;
    height: 0px;
    boder-radius:6px;
}
QScrollBar::handle:vertical{
background:black;
width:12px;
height:24px;
border-radius:6px;
min-height:24px;
max-height:24px;
border:1px solid white;
}
QScrollBar::groove:vertical{
background:#2c2f36;
width:12px;
border:0px solid black;
boder-radius:6px;
}
QScrollBar::handle:horizontal{
background:#585a5c;
width:12px;
border:0px solid black;
}
QScrollBar::groove:horizontal{
background:#3d3f42;
width:12px;
border:0px solid black;
}
'''
q=s
hola_je=s
class SimpleHighlighter(QSyntaxHighlighter):
    def __init__(self, parent=None):
        super().__init__(parent)

        # Text format for quoted text (green)
        self.greenFormat = QTextCharFormat()
        self.greenFormat.setForeground(QColor(91, 212, 107))  # Green for quoted text

        # Text format for comments (grey)
        self.greyFormat = QTextCharFormat()
        self.greyFormat.setForeground(QColor(169, 169, 169))  # Grey for comments

        # Text format for special elements
        self.orangeFormat = QTextCharFormat()
        self.orangeFormat.setForeground(QColor(255, 165, 0))

        self.yellowFormat = QTextCharFormat()
        self.yellowFormat.setForeground(QColor(255, 255, 0))  # Yellow for 'display'

        self.redFormat = QTextCharFormat()
        self.redFormat.setForeground(QColor(255, 0, 0))  # Red for errors
        
        # Regular expressions for highlighting rules
        self.singleQuoteRegex = QRegularExpression(r"'(.*?)'")
        self.doubleQuoteRegex = QRegularExpression(r'"(.*?)"')
        self.commentRegex = QRegularExpression(r'##.*')
        self.questionMarkRegex = QRegularExpression(r'\?')
        self.greaterThanRegex = QRegularExpression(r"^>.*")
        self.errorRegex = QRegularExpression(r'(#ERROR:)([^>]*)(>)?')

    def highlightBlock(self, text):
        # Match text starting with '>'
        match = self.greaterThanRegex.match(text)
        if match.hasMatch():
            start = match.capturedStart()
            self.setFormat(start, len(text) - start, self.orangeFormat)
        
        # Match and highlight text inside single quotes
        match = self.singleQuoteRegex.match(text)
        while match.hasMatch():
            start = match.capturedStart(1)
            length = match.capturedLength(1)
            self.setFormat(start, length, self.greenFormat)
            match = self.singleQuoteRegex.match(text, match.capturedEnd())

        # Highlight comments (text starting with '##')
        commentMatch = self.commentRegex.match(text)
        if commentMatch.hasMatch():
            self.setFormat(0, len(text), self.greyFormat)
        
        # Match and highlight text inside double quotes
        match = self.doubleQuoteRegex.match(text)
        while match.hasMatch():
            start = match.capturedStart(1)
            length = match.capturedLength(1)
            self.setFormat(start, length, self.greenFormat)
            match = self.doubleQuoteRegex.match(text, match.capturedEnd())
        
        # Highlight question marks
        match = self.questionMarkRegex.match(text)
        while match.hasMatch():
            start = match.capturedStart()
            self.setFormat(start, 1, self.orangeFormat)
            match = self.questionMarkRegex.match(text, match.capturedEnd())
        
        # Highlight #ERROR: in red and text after it in red until '>'
        match = self.errorRegex.match(text)
        if match.hasMatch():
            error_start = match.capturedStart(1)
            error_length = match.capturedLength(1)
            text_start = match.capturedStart(2)
            text_length = match.capturedLength(2)
            
            self.setFormat(error_start, error_length, self.redFormat)  # Highlight #ERROR:
            self.setFormat(text_start, text_length, self.redFormat)  # Highlight text after #ERROR:
        
        # Highlight "expr:" in orange
        expr_regex = QRegularExpression(r'\bexpr:')
        match_iter = expr_regex.globalMatch(text)
        while match_iter.hasNext():
            match = match_iter.next()
            start = match.capturedStart()
            length = match.capturedLength()
            self.setFormat(start, length, self.orangeFormat)

        # Highlight "display" in yellow
        display_regex = QRegularExpression(r'\bdisplay\b')
        match_iter = display_regex.globalMatch(text)
        while match_iter.hasNext():
            match = match_iter.next()
            start = match.capturedStart()
            length = match.capturedLength()
            self.setFormat(start, length, self.yellowFormat)

        # Highlight parentheses '(' and ')' in green
        paren_regex = QRegularExpression(r'[\(\)]')
        match_iter = paren_regex.globalMatch(text)
        while match_iter.hasNext():
            match = match_iter.next()
            start = match.capturedStart()
            self.setFormat(start, 1, self.greenFormat)

        read_docex_regex = QRegularExpression(r'\bread\.docex\b')
        match_iter = read_docex_regex.globalMatch(text)
        while match_iter.hasNext():
            match = match_iter.next()
            start = match.capturedStart()
            length = match.capturedLength()
            self.setFormat(start, length, self.yellowFormat)

        Variable_Retrieval()


def open_file_explorer():
    file_path,_=QFileDialog.getOpenFileName(None,"Select the file")
    if file_path:
        file_s,file_extension=os.path.splitext(file_path)
        if file_extension in ['.xlsx','.xlsm','.xltx','.xltm']:
            g=L.retrieve_from_excel(file_path)
            data_viewer.setHtml(g)
        else:
            data_viewer.setHtml("<body style='background-color:#2e2b2b;'><p style='color:red;background-color:#2e2b2b;'>Only xlsx,xlsm,xltx,xltm are allowed</p></body>")
    else:
        data_viewer.setHtml("<body style='background-color:#2e2b2b;'><p style='color:red;'>PLEASE SELECT A FILE</p></body>")

def Variable_Retrieval():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    g=L.retrieve_from_excel_variable_values(file_path)
    memory_section.setHtml(g)
    

from PyQt5.QtCore import Qt
import math
import sys
import os
from PyQt5.QtWidgets import QWidget,QInputDialog,QMessageBox,QVBoxLayout,QPushButton,QPlainTextEdit,QApplication
from PyQt5.QtCore import QDir

class file_explorer(QWidget):
    def __init__(self):
        super().__init__()
        self.layout=QVBoxLayout()
        self.INITui()
        self.history=[]
        self.setGeometry(400,400,900,900)
        self.current_path=QDir.rootPath()
        self.load_directory(self.current_path)

    def INITui(self):
        self.back_button=QPushButton("⬅ Back")
        self.back_button.clicked.connect(self.go_back)
        self.back_button.setEnabled(False)
        
        self.create_button=QPushButton("📂 Create New Folder")
        self.create_button.clicked.connect(self.create_folder)

        self.delete_folder=QPushButton("🗑 Delete Selected Item")
        self.delete_folder.clicked.connect(self.delete)

        self.text_edit=QPlainTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setStyleSheet(s+"font-family: Consolas; font-size: 14px;background-color:#2e2b2b;color: white; ")

        self.text_edit.mouseDoubleClickEvent=self.on_item_click

        self.layout.addWidget(self.create_button)
        self.layout.addWidget(self.delete_folder)
        self.layout.addWidget(self.back_button)
        self.layout.addWidget(self.text_edit)

        self.setLayout(self.layout)
    def on_item_click(self,event):
        cursor=self.text_edit.textCursor()
        cursor.select(cursor.LineUnderCursor)
        selected_item=cursor.selectedText().strip()
        if selected_item.startswith("📁 "):
            folder_name=selected_item.replace("📁 ", "")
            new_path=os.path.join(self.current_path,folder_name)
            self.history.append(self.current_path)
            self.load_directory(new_path)
        elif selected_item.startswith("📄 "):
            file_name=selected_item.replace("📄 ", "")
            file_path=os.path.join(self.current_path,file_name)
            global file_tab
            text=file_tab.currentWidget()
            file_tab.setTabText(file_tab.currentIndex(),selected_item)
            with open(file_path,'r') as file:
                content=file.read()
                text.setPlainText(content)
    def create_folder(self):
        folder_name,ok=QInputDialog.getText(self,"Create Folder","Enter new folder name")
        if ok and folder_name:
            new_folder_path=os.path.join(self.current_path,folder_name)
            try:
                os.makedirs(new_folder_path)
                QMessageBox.information(self,"Success",f"The folder {folder_name} has been created.")
                self.load_directory(new_folder_path)
            except Exception as e:
                QMessageDialog.critical(self,"Error!",f"The folder {folder_name} couldn't be created \n str(e)")
    def delete(self):
        cursor=self.text_edit.textCursor()
        cursor.select(cursor.LineUnderCursor)
        selected_item=cursor.selectedText().strip()
        if selected_item.startswith("📁 "):
            folder_name=selected_item.replace("📁 ", "")
            item_path=os.path.join(self.current_path,folder_name)
            reply = QMessageBox.question(self, "Delete Folder", f"Are you sure you want to delete {folder_name}?", 
                             QMessageBox.Yes | QMessageBox.No)
            
            if reply==QMessageBox.Yes:
                try:
                    os.rmdir(item_path)
                    self.load_directory(self.current_path)
                except Exception as e:
                    QMessageBox.critical(self,"Error",f"The folder {folder_name} couldn't be deleted \n {str(e)}")
        elif selected_item.startswith("📄 "):
            file_name=selected_item.replace("📄 ", "")
            file_path=os.path.join(self.current_path,file_name)
            reply = QMessageBox.question(self, 
                             f"Do you want to delete the file {file_name}?", 
                             "Are you sure you want to delete this file?", 
                             QMessageBox.Yes | QMessageBox.No, 
                             QMessageBox.No)
            
            if reply==QMessageBox.Yes:
                try:
                    os.remove(file_path)
                    QMessageBox.information(self,"Success! The file {file_name} has been removed!")
                    self.load_directory(current_path)
                except Exception as e:
                    QMessageBox.critical(self,f"Error! The file {file_name} couldn't be deleted.\n str(e)")
    def go_back(self):
        if self.history:
            previous_path=self.history.pop()
            self.load_directory(previous_path)
    def load_directory(self,path):
        self.text_edit.clear()
        self.text_edit.appendPlainText(f"Contents of :{path}")
        self.text_edit.appendPlainText("-"*50)
        directory=QDir(path)
        directory.setFilter(QDir.Dirs | QDir.NoDotAndDotDot | QDir.Files)
        items=directory.entryList()
        if items:
            for item in items:
                item_path=os.path.join(path,item)
                if os.path.isdir(item_path):
                    self.text_edit.appendPlainText(f"📁 {item}")
                elif item.endswith(".txt"):
                    self.text_edit.appendPlainText(f"📄 {item}")
        else:
            self.text_edit.appendPlainText("Empty Directory")
        self.current_path=path
        self.back_button.setEnabled(len(self.history)>0)





def tab_apply_styles(tab):
    tab.setStyleSheet("""
        QTabBar::tab {
            background-color: #2e2b2b;
            color: white;
        }
        QTabBar::tab:selected {
            background-color: #505057;
            color: white;
            
        }
    """)

def tab_apply_styles_2(tab):
    tab.setStyleSheet("""
        QTabBar::tab {
            background-color: #505057;
            color: white;
        }
        QTabBar::tab:selected {
            background-color: #2e2b2b;
            color: white;
            
        }
    """)
app=QApplication([])
central_widget=QWidget()
main_window=QMainWindow()
main_window.setStyleSheet("background-color:#585b6e;")
#close_first_instance()
main_window.setCentralWidget(central_widget)
desktop_dimensions=QDesktopWidget().screenGeometry()
width=desktop_dimensions.width()
height=desktop_dimensions.height()
main_window.setGeometry(0,0,width,height)

##layouts and design
main_layout=QHBoxLayout()
sublayout_1=QVBoxLayout()

text_editor=QPlainTextEdit()
highlighter=SimpleHighlighter(text_editor.document())
console=L.CustomPlainTextEdit("")


highlighter_2=SimpleHighlighter(console.document())
others=file_explorer()
others.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)


s += "QPlainTextEdit { background-color: #2e2b2b; color: white; }"
q=q+"QTextEdit { background-color: #2e2b2b; color: white; }"
text_editor.setStyleSheet(s)
console.setStyleSheet(q)
sublayout_2=QVBoxLayout()
memory_section=QTextEdit()
##plus_tab=QPlainTextEdit()

diff_files=QTabWidget()
#tab_apply_styles()
diff_files.setElideMode(Qt.ElideNone)
file_tab=QTabWidget()
file_tab.setContentsMargins(0,0,0,0)
text_editor.setLineWrapMode(QPlainTextEdit.NoWrap)
file_tab.addTab(text_editor,"Untitled")
##file_tab.addTab(plus_tab,"+")
tab_apply_styles_2(file_tab)
accomodation_file_section=QFrame()
accomodation_file_section.setStyleSheet("background-color:#313136;")
acc_layout=QHBoxLayout()
acc_layout.setContentsMargins(0,0,0,0)
acc_layout.addWidget(file_tab)
accomodation_file_section.setContentsMargins(0,0,0,0)
accomodation_file_section.setStyleSheet("background-color:#505057;")

accomodation_file_section.setLayout(acc_layout)
diff_files.addTab(accomodation_file_section,"Files")
tab_apply_styles(diff_files)
data_viewer_ultimate=QFrame()
data_viewer_ultimate.setStyleSheet("margin:0px;padding:0px;background-color:#3a3b3d;border:0px solid black;")

data_viewer_control=QHBoxLayout()

######file accesss#################
ml=QPushButton("File")
ml.clicked.connect(open_file_explorer)
#m.clicked.connect(None)
data_viewer_control.addWidget(ml,stretch=1)
data_viewer=QTextEdit()
ml.setStyleSheet("background-color:grey;")
data_viewer.setHtml("<body style='background-color:#2e2b2b;'></body>")
data_viewer.setReadOnly(True)
data_viewer_control.addWidget(data_viewer,stretch=13)
##data_viewer.setLineWrapMode(QPlainTextEdit.NoWrap)
data_viewer.setStyleSheet("background-color:white;")
data_viewer_control.setContentsMargins(5,0,0,0)
data_viewer_ultimate.setLayout(data_viewer_control)
diff_files.addTab(data_viewer_ultimate,"Data Viewer")
m=QFrame()
layout=QHBoxLayout()
layout.addWidget(diff_files)
layout.setContentsMargins(0,0,0,0)
m.setLayout(layout)
m.setStyleSheet("background-color:#2e2b2b")
sublayout_1.addWidget(m,stretch=50)
sublayout_1.addWidget(console,stretch=50)
s1=s
s1 += "QTextEdit { background-color: #2e2b2b; color: white;}"
memory_section.setStyleSheet(s1)
others.setStyleSheet("background-color:#2e2b2b;color:white;")
tab_apply_styles(diff_files)
sublayout_2.addWidget(memory_section,stretch=50)
sublayout_2.addWidget(others,stretch=50)

#m.setLayout(sublayout_1)
#m.setStyleSheet("background-color:black;")
main_layout.addLayout(sublayout_1,stretch=60)
main_layout.addLayout(sublayout_2,stretch=40)

##layouts and design


script_dir = os.path.dirname(os.path.abspath(__file__))
syntax_file_path = os.path.join(script_dir, 'Syntax_LaplaceLab.txt')
with open(syntax_file_path,'r') as file:
        content=file.read()
        text_editor.setPlainText(content)

def New():
    global file_tab
    new_Tab=QPlainTextEdit()
    new_Tab.setStyleSheet("background-color:#2e2b2b;color:white;")
    high_2=SimpleHighlighter(new_Tab.document())
    file_tab.addTab(new_Tab,"Untitled")
    
    
    
def Open():
    file_path,_=QFileDialog.getOpenFileName(None,"Select the file")
    if file_path:
        global file_tab
        file_s,file_extension=os.path.splitext(file_path)
        file_name=os.path.basename(file_path)
        current_text_editor=file_tab.currentWidget()
        if file_extension in ['.txt']:
            with open(file_path,'r') as file:
                content=file.read()
                current_text_editor=file_tab.currentWidget()
                file_tab.setTabText(file_tab.currentIndex(),file_name)
                current_text_editor.setPlainText(content)
        else:
            current_text_editor.setPlainText("Only .txt files are allowed")
    else:
        current_text_editor.setPlainText("You haven't chosen a file")

    
def Save():
    global file_tab
    current_text_editor=file_tab.currentWidget()
    extracted_text_editor=current_text_editor.toPlainText()
    if file_tab.tabText(file_tab.currentIndex())=="Untitled":
        file_path, _ = QFileDialog.getSaveFileName(None,"Save File", "", "Text Files (*.txt);;All Files (*)")
        file_name=os.path.basename(file_path)
        file_tab.setTabText(file_tab.currentIndex(),file_name)
        with open(file_path,'w') as file:
            file.write(extracted_text_editor)
    else:
        with open(file_tab.tabText(file_tab.currentIndex()),'w') as file:
            file.write(extracted_text_editor)
def Save_As():
     global file_tab
     current_text_editor=file_tab.currentWidget()
     extracted_text_editor=current_text_editor.toPlainText()
     file_path, _ = QFileDialog.getSaveFileName(None,"Save File", "", "Text Files (*.txt);;All Files (*)")
     file_name=os.path.basename(file_path)
     file_tab.setTabText(file_tab.currentIndex(),file_name)
     with open(file_path,'w') as file:
        file.write(extracted_text_editor)
def Exit():
    import openpyxl
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    wb=openpyxl.load_workbook(file_path)
    sheet=wb.active
    for row in sheet.iter_rows():
        for cell in row:
            cell.value=None
    wb.save(file_path)
    print("variable file emptied")
    sys.exit()
def Cut():
    text_edit_new=current_text_editor=file_tab.currentWidget()
    text_edit_new.cut()
def Copy():
    text_edit_new=current_text_editor=file_tab.currentWidget()
    text_edit_new.copy()
def Paste():
    text_edit_new=current_text_editor=file_tab.currentWidget()
    text_edit_new.paste()
def Zoom_In():
    text_edit_new=current_text_editor=file_tab.currentWidget()
    font=text_edit_new.font()
    font_size=font.pointSize()
    print(font_size)
    font_size=font_size+3
    font.setPointSize(font_size)
    text_edit_new.setFont(font)
def Zoom_Out():
    text_edit_new=current_text_editor=file_tab.currentWidget()
    font=text_edit_new.font()
    font_size=font.pointSize()
    font_size=font_size-3
    font.setPointSize(font_size)
    text_edit_new.setFont(font)
def Reset_Zoom():
    text_edit_new=current_text_editor=file_tab.currentWidget()
    font=text_edit_new.font()
    font_size=14
    font.setPointSize(font_size)
    text_edit_new.setFont(font)
def Equation_Plotter():
    threading.Thread(target=L.plotter).start()
def Documentation():
    import webbrowser
    file_path = "/home/suraj/Desktop/Laplacelab/laplacecode/Laplace_documentation.html"
    webbrowser.open(file_path)
def About_LaplaceLab():
    import webbrowser
    file_path = "/home/suraj/Desktop/Laplacelab/laplacecode/about_laplacelab.html"
    webbrowser.open(file_path)
import re

def custom_parser(input_text):
    input_text = re.sub(r'([a-zA-Z_][a-zA-Z0-9_]*)\?([\'"][^\'"]*[\'"]|[0-9]+)', r'\1=\2', input_text)
    return input_text

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
def insert_already_existing_excel_variable(item_name,value_var):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    wb=load_workbook(file_path);
    ws=wb.active;i=1;
    while ws['A'+str(i)].value!=None:
        if ws['A'+str(i)].value==item_name:
            print(ws['A'+str(i)])
            q=ws['B'+str(i)].value;
            ws['B'+str(i)]=value_var;
            wb.save(file_path)
            return 1   
        else:
            i=i+1;

def display(value):
    from Laplacelab import extract_from_excel_variable_list
    m=extract_from_excel_variable_list(value)
    global console_log
    console_log=console_log+"\n"+value+"\n"
operators = ['+', '-', '*', '/','?>','<?','??','!?','<','>']
def replace_variables_with_values_in_expr(m, expr):
    i = 0
    try:
        while i < len(expr):
            if expr[i] not in operators:
                if expr[i].isdigit():
                    pass  # Ignore numeric values
                else:
                    q = L.cmd_retrieve_excel(expr[i])
                    if q == "#ERROR":
                        console.append(">" + m)
                        console.append(f"#ERROR: Variable '{expr[i]}' doesn't exist\n")
                        return "#ERROR"  # Return early if any variable is missing
                    else:
                        expr[i] = q  # Replace variable with its value
            i += 1
    except Exception as e:
        print(e)
    
    try:
        print(expr)
        result = ''.join(map(str, expr))
        result = result.replace("??", "==").replace("!?", "!=").replace("<?", "<=").replace("?>", ">=")
        result = eval(result)
        print(result)
        return result
    except Exception as e:
        console.append(f"#ERROR: Invalid expression in '{m}'\n")
        return "#ERROR"
def evaluate_expression(expr):
    processed_expr = []
    
    for token in expr:
        if token.isalpha():  # Check if it's a variable (letters only)
            value = L.cmd_retrieve_excel(token)  # Retrieve value from Excel
            if value == "#ERROR":
                console.append(f"#ERROR: Variable '{token}' doesn't exist\n")
                return False  # Return False if a variable is missing
            processed_expr.append(str(value))  # Convert retrieved value to string
        else:
            processed_expr.append(token)  # Keep numbers and operators unchanged

    # Convert list to a string expression
    expression_str = ''.join(processed_expr)

    # Replace custom operators with Python equivalents
    replacements = {
        "??": "==",
        "!?": "!=",
        "<?": "<=",
        "?>": ">="
    }
    for key, value in replacements.items():
        expression_str = expression_str.replace(key, value)

    try:
        # Check for allowed characters to prevent unsafe execution
        allowed_chars = set("0123456789+-*/<>=!(). ")
        if all(c in allowed_chars for c in expression_str):
            return eval(expression_str)  # Evaluate and return True/False
        else:
            console.append(f"#ERROR: Invalid characters in expression '{expression_str}'\n")
            return False
    except Exception as e:
        console.append(f"#ERROR: Invalid expression '{expression_str}'\n")
        return False

def extract_inside_parentheses(s):
    start = s.find("(")
    end = s.find(")")
    
    if start != -1 and end != -1 and start < end:
        return s[start + 1:end]  
    return None
def extract_inside_braces(s):
    """Extracts the content inside the outermost { ... } while handling nested braces correctly."""
    start = s.find("{")
    if start == -1:
        return None  # No opening brace found
    
    stack = []
    for i in range(start, len(s)):
        if s[i] == "{":
            stack.append(i)  # Push to stack when encountering '{'
        elif s[i] == "}":
            stack.pop()  # Pop from stack when encountering '}'
            if not stack:  # If stack is empty, we've found the matching '}'
                return s[start + 1:i]  # Extract content properly

    return None  # Return None if no matching '}' is found

def is_quoted(s):
    return (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'"))
def remove_quotes(s):
    if len(s) >= 2 and (s[0] == s[-1]) and (s[0] in ['"', "'"]):
        return s[1:-1]  
    return s  
def split_statements(code):
    """Splits code into statements, keeping nested { ... } blocks intact."""
    statements = []
    current_statement = ""
    inside_braces = 0

    i = 0
    while i < len(code):
        char = code[i]
        current_statement += char

        if char == "{":
            inside_braces += 1
        elif char == "}":
            inside_braces -= 1

        if char == ";" and inside_braces == 0:
            statements.append(current_statement.strip().rstrip(";"))  # Remove trailing ";"
            current_statement = ""

        i += 1  # Ensure loop progresses

    if current_statement.strip():
        statements.append(current_statement.strip())  # Add remaining statement

    return statements

def compare_eval(expression):
    
    expression = expression.replace("<?", "<=")  
    expression = expression.replace(">?", ">=")  
    expression = expression.replace("?", "==")   
    expression = expression.replace("!=", "!?") 
    return eval(expression)
def preserve_spaces(user_code):
    result = []
    inside_comment = False
    inside_string = False
    temp_comment = ""
    temp_string = ""
    
    i = 0
    while i < len(user_code):
        char = user_code[i]
        
        # Detect start of a comment (##)
        if char == "#" and i + 1 < len(user_code) and user_code[i + 1] == "#":
            inside_comment = True
            temp_comment = "##"
            i += 1  # Skip next '#'
        elif inside_comment:
            temp_comment += char
            if char == ";":  # End of comment
                inside_comment = False
                result.append(temp_comment)
                temp_comment = ""
        # Detect start of a string (")
        elif char == '"' and not inside_comment:
            if inside_string:
                inside_string = False
                temp_string += char
                result.append(temp_string)
                temp_string = ""
            else:
                inside_string = True
                temp_string = char
        elif inside_string:
            temp_string += char
        else:
            # Only add non-space characters outside strings and comments
            if char != " ":
                result.append(char)

        i += 1
    
    return "".join(result)
import sys
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QDialog
import numpy as np
import ast
class PlotWindow(QDialog):
    def __init__(self,variables,graph_name,m):
        super().__init__()
        self.setWindowTitle("Diagram")
        self.setGeometry(500,500,500,500)
        self.canvas=FigureCanvas(plt.figure())
        layout=QVBoxLayout()
        layout.addWidget(self.canvas)
        self.plot_graph(variables,graph_name)
        self.setLayout(layout)
        try:
            self.plot_graph(variables,graph_name)
        except Exception as e:
            console.append(">"+m)
            console.append(str(e))
    def plot_graph(self,variables,graph_name):
      if graph_name=="scatter_plot":
        try:
            import ast
            import matplotlib.pyplot as plt
            import numpy as np
            x_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[0]))
            y_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[1]))
            x_axis=np.array(x_axis)
            y_axis=np.array(y_axis)
            x_label=variables[2]
            y_label=variables[3]
            fig,ax=plt.subplots(figsize=(6,4))
            ax.scatter(x_axis,y_axis,color="blue",alpha=0.7)
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.set_title("Scatter Plot")
            ax.grid()
            self.canvas.figure=fig
            self.canvas.draw
        except Exception as e:
            print(str(e))
      elif graph_name=="bar_chart":
            import ast
            import matplotlib.pyplot as plt
            import numpy as np
            x_axis = [str(x) for x in ast.literal_eval(L.cmd_retrieve_excel(variables[0]))]  
            y_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[1]))
            x_label=variables[2]
            y_label=variables[3]
            fig,ax=plt.subplots(figsize=(6,4))
            ax.bar(x_axis,y_axis,color="blue")
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.set_title("Bar chart")
            ax.grid()
            self.canvas.figure=fig
            self.canvas.draw
      elif graph_name=="pie_chart":
            import ast
            import matplotlib.pyplot as plt
            import numpy as np
            x_axis = [str(x) for x in ast.literal_eval(L.cmd_retrieve_excel(variables[0]))]  
            y_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[1]))
            x_label=variables[2]
            y_label=variables[3]
            fig,ax=plt.subplots(figsize=(6,4))
            colors = plt.get_cmap("tab10")(np.linspace(0, 1, len(x_axis)))
            ax.pie(y_axis,labels=x_axis,autopct="%1.1f%%",colors=colors)
            ax.set_title("Pie chart")
            ax.grid()
            self.canvas.figure=fig
            self.canvas.draw
      elif graph_name=="line_graph":
            import ast
            import matplotlib.pyplot as plt
            import numpy as np
            x_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[0]))  
            y_axis=ast.literal_eval(L.cmd_retrieve_excel(variables[1]))
            x_label=variables[2]
            y_label=variables[3]
            fig,ax=plt.subplots(figsize=(6,4))
            ax.plot(x_axis,y_axis,marker="o",linestyle="-",color="blue",label="Line")
            ax.set_xlabel(x_label)
            ax.set_ylabel(y_label)
            ax.set_title("Basic Line Graph")
            ax.legend()
            ax.grid()
            
            self.canvas.figure = fig
            self.canvas.draw() 
      else:
            console.append(">"+m)
            console.append("#ERROR:PLEASE CHOOSE AN APPROPRIATE GRAPH TYPE.")
            return 0



def Run(c=None):
    import openpyxl
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    wb=openpyxl.load_workbook(file_path)
    sheet=wb.active
    if c is None or not isinstance(c, str):
       for row in sheet.iter_rows():
           for cell in row:
               cell.value=None
       wb.save(file_path)
       current_text_editor=file_tab.currentWidget()
       user_code = current_text_editor.toPlainText()
    else:
        user_code=c
    user_code=user_code.replace("\n","")
    user_code=preserve_spaces(user_code)
    user_code=split_statements(user_code)
    global console
    print(user_code)
    i=0
    while i <len(user_code):
     m=user_code[i]
     try:
        hi=0
        if m.startswith("plot."):
            new_m=m.removeprefix("plot.")
            print(new_m)
            print("hello")
            if new_m.startswith("scatter_plot"):
                variables=extract_inside_parentheses(new_m)
                print(variables)
                variables=variables.split(',')
                print(variables)
                if not variables:
                   console.append(">"+m)
                   addition=" Syntax doesnt match"
                   console.append(addition)
                elif len(variables)!=4:
                    console.append(">"+m)
                    addition=" Scatter plot needs two parameters"
                    console.append(addition)
                else:
                    i=0
                    print("ham")
                    while i<2:
                        print(variables[i])
                        print(L.check_excel_variable(variables[i]))
                        if not L.check_excel_variable(variables[i]):
                            console.append(">"+m)
                            addition=f" #ERROR:The variable doesnt exist\n"  
                            console.append(addition,"scatter_plot",m)
                            return 0
                        i=i+1
                    try:
                        qs=PlotWindow(variables,"scatter_plot",m)
                        qs.setGeometry(1000,1000,1000,1000)
                        qs.exec_()
                        return
                    except Exception as e:
                        console.append(">"+m)
                        addition=" #ERROR:"+str(e)
                        console.append(addition)
            elif new_m.startswith("bar_chart"):
                variables=extract_inside_parentheses(new_m)
                print(variables)
                variables=variables.split(',')
                print(variables)
                if not variables:
                   console.append(">"+m)
                   addition=" Syntax doesnt match"
                   console.append(addition)
                elif len(variables)!=4:
                    console.append(">"+m)
                    addition=" Scatter plot needs two parameters"
                    console.append(addition)
                else:
                    i=0
                    print("ham")
                    while i<2:
                        print(variables[i])
                        print(L.check_excel_variable(variables[i]))
                        if not L.check_excel_variable(variables[i]):
                            console.append(">"+m)
                            addition=f" #ERROR:The variable doesnt exist\n"  
                            console.append(addition)
                            return 0
                        i=i+1
                    try:
                        qs=PlotWindow(variables,"bar_chart",m)
                        qs.setGeometry(1000,1000,1000,1000)
                        qs.exec_()
                        return
                    except Exception as e:
                        console.append(">"+m)
                        addition=" #ERROR:"+str(e)
                        console.append(addition)
            elif new_m.startswith("pie_chart"):
                variables=extract_inside_parentheses(new_m)
                print(variables)
                variables=variables.split(',')
                print(variables)
                if not variables:
                   console.append(">"+m)
                   addition=" Syntax doesnt match"
                   console.append(addition)
                elif len(variables)!=4:
                    console.append(">"+m)
                    addition=" Pie chart needs two parameters"
                    console.append(addition)
                else:
                    i=0
                    print("ham")
                    while i<2:
                        print(variables[i])
                        print(L.check_excel_variable(variables[i]))
                        if not L.check_excel_variable(variables[i]):
                            console.append(">"+m)
                            addition=f" #ERROR:The variable doesnt exist\n"  
                            console.append(addition)
                            return 0
                        i=i+1
                    try:
                        qs=PlotWindow(variables,"pie_chart",m)
                        qs.setGeometry(1000,1000,1000,1000)
                        qs.exec_()
                        return
                    except Exception as e:
                        console.append(">"+m)
                        addition=" #ERROR:"+str(e)
                        console.append(addition)
            elif new_m.startswith("Line_Graph"):
                variables=extract_inside_parentheses(new_m)
                print(variables)
                variables=variables.split(',')
                print(variables)
                if not variables:
                   console.append(">"+m)
                   addition=" Syntax doesnt match"
                   console.append(addition)
                elif len(variables)!=4:
                    console.append(">"+m)
                    addition=" Line graph needs two parameters"
                    console.append(addition)
                else:
                    i=0
                    print("ham")
                    while i<2:
                        print(variables[i])
                        print(L.check_excel_variable(variables[i]))
                        if not L.check_excel_variable(variables[i]):
                            console.append(">"+m)
                            addition=f" #ERROR:The variable doesnt exist\n"  
                            console.append(addition)
                            return 0
                        i=i+1
                    try:
                        qs=PlotWindow(variables,"line_graph",m)
                        qs.setGeometry(1000,1000,1000,1000)
                        qs.exec_()
                        return
                    except Exception as e:
                        console.append(">"+m)
                        addition=" #ERROR:"+str(e)
                        console.append(addition)
        elif m.startswith("##"):
          pass  
        elif m.startswith("loop("):
            trial=extract_inside_parentheses(m)
            trial=L.cmd_retrieve_excel(trial)
            code=extract_inside_braces(m)
            i1=0
            while i1<int(trial):
                Run(code)
                i1=i1+1
        elif m.startswith("if("):
            condition=extract_inside_parentheses(m)
            condition=L.tokenize_expression(condition)
            print(condition)
            #print(replace_variables_with_values_in_expr(m,list_of_values))
            if condition=="":
                addition="\nSyntax Doesn't match\n"
                console.append(addition)
                return 0
            elif replace_variables_with_values_in_expr(m,condition):
              
              #print(replace_variables_with_values_in_expr(m,list_of_values))
              code=extract_inside_braces(m)
              print(code)
              print("hello")
              Run(code)
              if i+1<len(m) and (user_code[i+1].startswith("else")):
                i=i+1
            else:
                
                if user_code[i+1].startswith("else"):
                   code=extract_inside_braces(user_code[i+1])
                   Run(code)
                   i=i+1
                
                
        elif m.startswith("read.docex"):
            path=extract_inside_parentheses(m)
            console.append(">"+m)
            addition=L.retrieve_from_excel_console(path)
            print(addition)
            console.append(addition)
        elif "display" in m:
            part_inside_brackets=extract_inside_parentheses(m)
            if part_inside_brackets=="":
                console.append(">"+m)
                addition=" Syntax doesnt match"
                console.append(addition)
            elif is_quoted(part_inside_brackets):
                to_be_printed=remove_quotes(part_inside_brackets)
                console.append(">"+m)
                addition=" "+to_be_printed  
                console.append(addition+"\n")
            elif L.check_excel_variable(part_inside_brackets):
                value_to_be_appended=L.cmd_retrieve_excel(part_inside_brackets)
                console.append(">"+m)
                addition=" "+str(value_to_be_appended)  
                console.append(addition+"\n")
            elif any(op in part_inside_brackets for op in operators) and part_inside_brackets.startswith("expr:"):
                part_inside_brackets=part_inside_brackets.removeprefix("expr:")
                list_of_values = L.tokenize_expression(part_inside_brackets)
                answer=replace_variables_with_values_in_expr(m,list_of_values)
                if answer=="#ERROR":
                    console.append(">"+m)
                    addition=f" The variable doesnt exist\n"  
                    console.append(addition)
                else:
                    console.append(">"+m)
                    addition=" "+str(answer)+"\n"
                    console.append(addition)
            else:
                console.append(">"+m)
                addition=f" The variable doesnt exist\n"  
                console.append(addition)
                
        elif '?' in m:
            arr=m.split('?')
            name=arr[0]
            value=arr[1]
            if any(c in name for c in "[]{}"):
                console.append(">"+m)
                addition=" Variable names can't have [,],{ or }.\n"  
                console.append(addition)        
            elif value.startswith('[') and value.endswith(']'):
                print(L.check_excel_variable(name))
                if L.check_excel_variable(name):
                    print("hello00")
                    L.replace_excel(name,value)
                    print("hello1")
                    console.append(">"+m)
                    addition=f" Initialized variable {name} with {value}.\n"  
                    console.append(addition)
                    print("hello2")
                    pass
                else:
                    from laplacelab_miscellaneous import insert_into_excel_variable_list
                    insert_into_excel_variable_list(name,value)
                    console.append(">"+m)
                    addition=f" Initialized variable {name} with {value}.\n"  
                    console.append(addition)
                    pass
            elif '[' in value:
                array=value.split("[")[0]
                if L.check_excel_variable(array):
                    import ast
                    variable_value=L.cmd_retrieve_excel(array)
                    variable_value=ast.literal_eval(variable_value)
                    print(variable_value)
                    index=value.split('[')[1]
                    print(index)
                    index=list(index)
                    print(index)
                    index.pop(len(index)-1)
                    print(index)
                    index=int(index[0])
                    print(index)
                    value_to_be=variable_value[index]
                    print(value_to_be)
                    if L.check_excel_variable(name):
                        L.replace_excel(name,value_to_be)
                        console.append(">"+m)
                        addition=f" Initialized variable {name} with {value_to_be}.\n"  
                        console.append(addition)
                    else:
                        from laplacelab_miscellaneous import insert_into_excel_variable_list
                        insert_into_excel_variable_list(name,value_to_be)
                        console.append(">"+m)
                        addition=f" Initialized variable {name} with {value}.\n"  
                        console.append(addition)
                else:
                    console.append(">"+m)
                    addition=f" Array with the name {array} doesn't exist.\n"  
                    console.append(addition)
                    
            elif not (re.search(r'"', value) or re.search(r"'", value)):
             if value.startswith("retrieve_column"):
                 array=extract_inside_parentheses(value)
                 print(array)
                 array=array.split(",")
                 print(array)
                 path=array[0]
                 column=array[1]
                 print("path is"+path+"column is"+column)
                 var_value=L.column_retrieve(path,column)
                 print(var_value)
                 print(L.check_excel_variable(name))
                 var_value=str(var_value)
                 if L.check_excel_variable(name):
                    L.replace_excel(name,var_value)
                    console.append(">"+m)
                    addition=f" Initialized variable {name} with {var_value}.\n"  
                    console.append(addition)
                 else:
                    from laplacelab_miscellaneous import insert_into_excel_variable_list
                    insert_into_excel_variable_list(name,var_value)
                    console.append(">"+m)
                    addition=f" Initialized variable {name} with {var_value}.\n"  
                    console.append(addition)
             elif any(op in value for op in operators):
                list_of_values = L.tokenize_expression(value)
                answer=replace_variables_with_values_in_expr(m,list_of_values)
                if answer=="#ERROR":
                    console.append(">"+m)
                    addition=f" The variable doesnt exist\n"  
                    console.append(addition)
                    h=1
                    value=answer
                elif L.check_excel_variable(name):
                    L.replace_excel(name,answer)
                    console.append(">"+m)
                    
                    addition=f" Initialized variable {name} with {answer}.\n"  
                    console.append(addition)
                    h=1
                    value=answer
                else:
                    from laplacelab_miscellaneous import insert_into_excel_variable_list
                    insert_into_excel_variable_list(name,answer)
                    console.append(">"+m)
                    addition=f" Initialized variable {name} with {answer}.\n"  
                    console.append(addition)
                    h=1
                    value=answer
             elif re.search(r"[a-zA-Z]", value) and h!=1:
                name_v=value
                value=L.cmd_retrieve_excel(value)
                if value=="#ERROR":
                    console.append(">"+m)
                    console.append(f"#ERROR:Variable '{name_v}'doesn't exist\n")
                else:
                    hi=1    
                    if L.check_excel_variable(name):
                        L.replace_excel(name,value)
                    else:
                        from laplacelab_miscellaneous import insert_into_excel_variable_list
                        insert_into_excel_variable_list(name,value)
                    console.append(">"+m)
                    addition=f" Initialized variable {name} with {value}.\n"  
                    console.append(addition)
            
             elif hi!=1:
              if insert_already_existing_excel_variable(name,value):
                L.replace_excel(name,value)
                console.append(">"+m)
                console.append(f"Initialized {name} with {value}\n")
              else:
                from laplacelab_miscellaneous import insert_into_excel_variable_list
                insert_into_excel_variable_list(name,value)
                console.append(">"+m)
                addition=f" Initialized variable {name} with {value}.\n"  
                console.append(addition)
            hi=0
            
        else:
            print(user_code)
            addition="\nSyntax Doesn't match\n"
            console.append(addition)
            
        i=i+1   
     except Exception as e:
        addition=f"Error: {e}"
        console=L.CustomPlainTextEdit(addition)

    



class HoverAction(QAction):
    def __init__(self,text,parent=None):
        super().__init__(text,parent)
        self.setStyleSheet("QAction:hover{background-color:lightblue;}")
menu_structure = {
    "File": ["New", "Open", "Save", "Save_As","Run","Exit"],
    "Edit": ["Cut", "Copy", "Paste"],
    "View": ["Zoom_In", "Zoom_Out", "Reset_Zoom"],
    "Tools": ["Equation_Plotter"],
    "Help": ["Documentation", "About_LaplaceLab"]
}
def create_action(name,parent,triggered_function):
    m=QAction(name,parent)
    m.triggered.connect(triggered_function)
    return(m)


def menu(parent,menu_dict):
    lst=list(menu_dict.keys())
    menu=parent.menuBar()
    i=0
    menu=parent.menuBar()
    menu.setStyleSheet("""
        QMenuBar {
            background-color: #2e3831;
            color: white;
        }
        QMenu::item {
            background-color: transparent;
            color: white;
            padding: 5px 10px;
        }
        QMenu::item:selected {
            background-color: #0339fc;
            color: white;
        }
        
    """)
    while i<len(lst):
        section=menu.addMenu(lst[i])
        r_1=lst[i]
        r=menu_dict[r_1]
        j=0
        while j<len(r):
          func_ref=globals().get(r[j])
          section.addAction(create_action(r[j],main_window,func_ref))
          j=j+1
        i=i+1

menu(main_window,menu_structure)
central_widget.setLayout(main_layout) 
#central_widget.setStyleSheet("background-color:black;")
main_window.show()
app.exec_()

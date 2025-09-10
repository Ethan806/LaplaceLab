import os

script_dir = os.path.dirname(os.path.abspath(__file__))
variables_xl_path=os.path.join(script_dir,"Variable_Storage.xlsx")
##name=input("enter the path:");##name2=input("input the name of the file");
##def showfiles(name):
##        try:
##            with os.scandir(name) as entries:
##                for entry in entries:
##                    print(entry.name)
##        except:
##            print("error");
##
##def searchfiles(name2):
##    global name;
##    try:
##            with os.scandir(name) as entries:
##                for entry in entries:
##                    if (name2==entry.name):
##                        print("Location:"+name+'/'+name2);
##    except:
##            print("error");
####symbol=input("Enter the direction:");
##def changepath(symbol):
##    global name
##    if (symbol=='./'):
##        k=name[::-1];
##        s=0;i=0;m='';
##        while (s!="/"):
##               s=k[i];
##               m=m+s;
##               i=i+1;
##        f=m[::-1];
##        name=name.replace(f,'');
##    if(symbol[0]=='/'):
##        name=name+symbol;
##searchfiles('TEMPLATE.A51');
##            

##import subprocess
##import sys
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter 
##wb=load_workbook(name);
##ws=wb.active;
##sheet=wb['Data entry station'];
##for row in range(1,40):
##    for col in range(1,4):
##        char=get_column_letter(col);
##        print(sheet[char+str(row)].value);
##        
##wb.save("hello.xlsx");

##def retrieve_from_excel(path,col_name):
##    wb=load_workbook(path);
##    ws=wb.active;
##    s=True;i=0;q=1;
##    while (s==True):
##        char=get_column_letter(q);
##        
##        if(ws[char+'1'].value==None):
##            s=False;break;
##        q=q+1;
##    for m in range (1,q+1):
##        char=get_column_letter(m);
##        if (col_name==ws[char+'1'].value):
##            al=char;
##    s=True;q=1;
##    while (s==True):
##        if(ws['A'+str(q)].value==None):
##            s=False;
##        else:
##            q=q+1;
##    f=[];
##    for i in range(3,q+1):
##        if (str(ws[al+str(i)].value).isalpha()==False):
##           f.append(ws[al+str(i)].value);
##    return(f)
##retrieve_from_excel(name,'T.Bonds')


##function to check the value of  aprticular variable stored in excel
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
def extract_from_excel_variable_list(item_name):
    wb=load_workbook(variables_xl_path);
    ws=wb.active;i=1;
    while ws['A'+str(i)].value!=None:
        if ws['A'+str(i)].value==item_name:
            q=ws['B'+str(i)].value;
            import ast
            array=ast.literal_eval(q);
            return(array)
        else:
            i=i+1;

##function into insert values into excel
def insert_into_excel_variable_list(item_name, item_value):
    import os
    from openpyxl import load_workbook

    row = 1
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path=os.path.join(script_dir,"Variable_Storage.xlsx")
    
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    
    # Find the next empty row in column A
    while worksheet['A' + str(row)].value is not None:
        row += 1
    
    # Insert the item_name and item_value into the next available row
    worksheet['A' + str(row)] = item_name
    worksheet['B' + str(row)] = item_value
    
    workbook.save(file_path)


    
##TERMINAL SYNTAX PROGRAMS

##THis followign command first divides the string into two parts.
#The first part is then searched in the excel file, which eseentially contaisn the path for the excel file of the data, while the second part fo the entry is the oclumn name

def navigation_col_xl(entry):
    
    if "=" in entry:
        b=entry.index['='];
        k=entry[:b];
        entry2=entry[b+1:]
        if "$" in entry2:
            indx=entry.index('$');
            xl=entry[:indx];
            q=entry[indx+1:];
            m=var_check(variables_xl_path,xl);##1st parameter is the path of the excelf ile while the second parameter is the ariable name we wish to find
            arr= retrieve_from_excel(m,q);
            arr=str(arr);
            insert_into_excel_variable_list(k,arr);
        elif "+" and"-" and "/" and "*" and "array" not in entry2 :
            insert_into_excel_variable_list(k,b);
            
        elif "+" and"-" and "/" and "*" in entry2:
            import sympy;
            insert_into_excel_variable_list(k,sympy.simpify(entry2))
        elif "array" in entry2:
            g=entry2.replace('array(','');
            g=g.replace(')','');
            g='['+g+']';
            insertinto_excel_variable_list(k,g);
        
            
            
            
